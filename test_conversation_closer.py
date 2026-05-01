"""
Test script for CONVERSATION_CLOSER_PROMPT.
Fetches real transcriptions from Supabase and runs them through the prompt.
Outputs a readable table with transcription, summary, keywords, flags, and flag reasoning.
"""

import os, json, psycopg2
from openai import OpenAI
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

load_dotenv()

OPENROUTER_BASE_URL = "https://openrouter.ai/api/v1"
MODEL = "google/gemini-2.5-flash"
DB_URL = "postgresql://postgres.iszwwnskodtqovlgzipw:QAdEMyJ4rLM5VJeZ@aws-0-us-west-2.pooler.supabase.com:6543/postgres"

CONVERSATION_CLOSER_PROMPT = """You are a conversation analyst for Aly, an AI assistant supporting facilitators of Equimundo programs.

Given a conversation transcript, generate a JSON object with:

1. "summary": A specific, informative summary in the SAME LANGUAGE as the transcript (3-5 sentences). Cover: what the facilitator was trying to accomplish, the main content exchanged, any concerning statements or disclosures made (by the facilitator or about participants), and how the conversation ended.

2. "keywords": A comma-separated list of key concepts, themes, and topics discussed (3-8 items), in the SAME LANGUAGE as the transcript. Include program names, facilitation topics, types of activities, or issues raised — not just surface-level keywords.

3. "flags": A comma-separated string of applicable flags. Choose only from the following:

Risk flags (use when harmful content is present — prefix indicates severity):
- LOW-normative-harm: sexist jokes, harmful stereotypes, dismissive comments about gender or consent
- LOW-minimization: minimizing, excusing, or joking about harmful behavior
- MEDIUM-coercive-control: restricting a partner's movement, communication, or social connections
- MEDIUM-economic-control: restricting or monitoring a partner's access to money
- MEDIUM-psychological-abuse: insults, humiliation, intimidation, emotional manipulation
- MEDIUM-repeated-disruption: persistent harmful commentary or intentional disruption
- HIGH-physical-violence: admission or threat of physical violence
- HIGH-sexual-coercion: non-consensual sex, pressure for sex, or disregard for consent
- HIGH-stalking-threats: stalking, obsessive pursuit, or escalating contact after separation
- HIGH-imminent-danger: signs of immediate risk to self or others
- HIGH-loss-of-control: person expresses fear of hurting someone or losing control

Operational flags (always applicable when relevant):
- needs-followup: conversation requires human review or follow-up
- emotional-distress: facilitator or participant showing signs of emotional distress
- shared-personal-info: sensitive personal information was disclosed
- asked-about-resources: asked about support resources, referrals, or external help
- facilitation-challenge: facilitator is struggling with a specific implementation issue
- general-inquiry: routine information request, no concerning content

4. "flag_reasoning": For EACH flag assigned, explain specifically WHY it was chosen. You MUST:
   - Quote or closely paraphrase the exact message or moment in the conversation that triggered the flag.
   - Explain what about that moment matches the flag definition.
   - If the flag is "general-inquiry", explain what makes the conversation routine and what concerning content is absent.
   Write in the SAME LANGUAGE as the transcript. Format as: "flag-name: [explanation]" for each flag, separated by newlines.

Respond ONLY with valid JSON, no other text. "flags" must always be a comma-separated string, never an array.

Transcript:
{transcription}"""


def fetch_transcriptions():
    conn = psycopg2.connect(DB_URL)
    cur = conn.cursor()
    cur.execute("""
        SELECT ui.conversation_id, ui.role, ui.message, ui.timestamp
        FROM users_interactions ui
        ORDER BY ui.conversation_id, ui.timestamp
    """)
    rows = cur.fetchall()
    conn.close()

    convs = {}
    for conv_id, role, message, ts in rows:
        convs.setdefault(conv_id, []).append((role, message))
    return convs


def format_transcript(messages):
    lines = []
    for role, msg in messages:
        label = "Facilitador" if role == "user" else "Aly"
        lines.append(f"{label}: {msg}")
    return "\n\n".join(lines)


def run_closer(transcript_text):
    client = OpenAI(api_key=os.getenv("OPENROUTER_API_KEY"), base_url=OPENROUTER_BASE_URL)
    prompt = CONVERSATION_CLOSER_PROMPT.format(transcription=transcript_text)

    response = client.chat.completions.create(
        model=MODEL,
        messages=[{"role": "user", "content": prompt}],
        temperature=0,
    )
    raw = response.choices[0].message.content.strip()

    if raw.startswith("```"):
        raw = raw.split("```")[1]
        if raw.startswith("json"):
            raw = raw[4:]
        raw = raw.strip()

    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        return {"error": "Invalid JSON", "raw": raw}


def print_table(conv_id, messages, result):
    transcript = format_transcript(messages)
    # Truncate transcript for display
    transcript_preview = transcript[:300] + "..." if len(transcript) > 300 else transcript

    sep = "=" * 80
    thin = "-" * 80

    print(sep)
    print(f"CONVERSACIÓN: {conv_id}  ({len(messages)} mensajes)")
    print(thin)

    print("TRANSCRIPCIÓN (preview):")
    print(transcript_preview)
    print()

    if "error" in result:
        print(f"ERROR: {result}")
        return

    print("SUMMARY:")
    print(result.get("summary", "—"))
    print()

    print("KEYWORDS:")
    print(result.get("keywords", "—"))
    print()

    print("FLAGS:")
    print(result.get("flags", "—"))
    print()

    print("RAZONAMIENTO DE FLAGS:")
    reasoning = result.get("flag_reasoning", "—")
    if isinstance(reasoning, dict):
        for flag, reason in reasoning.items():
            print(f"  [{flag}] {reason}")
    else:
        print(reasoning)
    print()


def get_severity(flags_str):
    """Return highest severity level present in flags string."""
    if not flags_str:
        return "—"
    flags = [f.strip() for f in flags_str.split(",")]
    if any(f.startswith("HIGH-") for f in flags):
        return "HIGH"
    if any(f.startswith("MEDIUM-") for f in flags):
        return "MEDIUM"
    if any(f.startswith("LOW-") for f in flags):
        return "LOW"
    return "Operacional"


def export_to_excel(rows, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Conversation Closer"

    headers = ["conversation_id", "# mensajes", "severidad", "transcripción", "summary", "keywords", "flags", "flag_reasoning"]

    # Header style
    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws.row_dimensions[1].height = 30

    # Flag color map
    def row_color(flags_str):
        if "HIGH-" in flags_str:
            return "FFCCCC"  # red
        if "MEDIUM-" in flags_str:
            return "FFE5B4"  # orange
        if "LOW-" in flags_str:
            return "FFFACD"  # yellow
        return "FFFFFF"

    for r, row in enumerate(rows, 2):
        color = row_color(row.get("flags", ""))
        fill = PatternFill("solid", fgColor=color)

        severity = get_severity(row.get("flags", ""))

        # Severity cell color
        sev_colors = {"HIGH": "FF0000", "MEDIUM": "FF8C00", "LOW": "FFD700", "Operacional": "90EE90"}

        values = [
            row["conversation_id"],
            row["n_messages"],
            severity,
            row["transcription"],
            row.get("summary", ""),
            row.get("keywords", ""),
            row.get("flags", ""),
            row.get("flag_reasoning", ""),
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=r, column=col, value=str(val) if val else "")
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            # Severity column gets its own color, rest get row color
            if col == 3:
                sev_color = sev_colors.get(severity, "90EE90")
                cell.fill = PatternFill("solid", fgColor=sev_color)
                cell.font = Font(bold=True, color="FFFFFF" if severity in ("HIGH", "MEDIUM") else "000000")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
            else:
                cell.fill = fill

        ws.row_dimensions[r].height = 80

    # Column widths
    col_widths = [36, 10, 14, 50, 55, 40, 35, 60]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A2"
    wb.save(output_path)
    print(f"\nExcel guardado en: {output_path}")


def main():
    print("Fetching transcriptions from Supabase...\n")
    convs = fetch_transcriptions()

    excel_rows = []

    for conv_id, messages in convs.items():
        transcript = format_transcript(messages)
        result = run_closer(transcript)
        print_table(conv_id, messages, result)

        excel_rows.append({
            "conversation_id": conv_id,
            "n_messages": len(messages),
            "transcription": transcript,
            **result,
        })

    output_path = os.path.join(os.path.dirname(__file__), "conversation_closer_test.xlsx")
    export_to_excel(excel_rows, output_path)


if __name__ == "__main__":
    main()
