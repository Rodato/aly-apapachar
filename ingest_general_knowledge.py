#!/usr/bin/env python3
"""
Ingesta de conocimiento general de Equimundo en MongoDB.
Colección destino: aly_general_knowledge (misma DB que apapachar)

Uso:
  python3 ingest_general_knowledge.py

Requiere: python-docx, openpyxl, pymongo, openai (via requests)
"""

import os
import time
import logging
import hashlib
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Optional

import numpy as np
import requests
from pymongo import MongoClient
from dotenv import load_dotenv

load_dotenv()
logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
DOCS_DIR = Path(__file__).parent / "data" / "docs" / "individual"
EXCEL_PATH = Path(__file__).parent / "data" / "ALY_Knowledge_Base.xlsx"
TARGET_COLLECTION = "aly_general_knowledge"
MAX_WORDS_PER_CHUNK = 300
OVERLAP_PARAGRAPHS = 1
EMBEDDING_DELAY = 0.1  # seconds between OpenAI calls


# ---------------------------------------------------------------------------
# Excel metadata loader
# ---------------------------------------------------------------------------
def load_metadata_index(xlsx_path: Path) -> Dict[str, Dict]:
    """
    Returns dict keyed by document_name (without .docx extension).
    Values: {knowledge_type, country, program, theme_category, language, keywords, status}
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("Run: pip install openpyxl")

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    headers = [str(cell.value).strip() if cell.value else "" for cell in next(ws.iter_rows())]
    logger.info(f"Excel headers: {headers}")

    index = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))
        doc_name = str(row_dict.get("document_name", "") or "").strip()
        if not doc_name or doc_name == "None":
            continue

        # Normalize theme_category to snake_case
        theme_raw = str(row_dict.get("theme_category", "") or "").strip()
        theme = _normalize_theme(theme_raw)

        # Parse keywords: comma-separated string → list
        kw_raw = str(row_dict.get("keywords", "") or "").strip()
        keywords = [k.strip() for k in kw_raw.split(",") if k.strip()] if kw_raw else []

        index[doc_name] = {
            "knowledge_type": str(row_dict.get("knowledge_type", "general") or "general").strip(),
            "country": str(row_dict.get("country", "") or "").strip(),
            "program": str(row_dict.get("program", "") or "").strip(),
            "theme_category": theme,
            "language": str(row_dict.get("language", "es") or "es").strip().lower(),
            "keywords": keywords,
            "status": str(row_dict.get("status", "Available") or "Available").strip(),
            "content_summary": str(row_dict.get("content_summary", "") or "").strip(),
        }
    logger.info(f"Loaded metadata for {len(index)} documents from Excel")
    return index


def _normalize_theme(theme_raw: str) -> str:
    mapping = {
        "marco teórico": "marco_teorico",
        "marco teorico": "marco_teorico",
        "tips facilitadores": "tips_facilitadores",
        "mejores prácticas": "mejores_practicas",
        "mejores practicas": "mejores_practicas",
        "rompehielos": "rompehielos",
        "biblioteca de recursos": "recursos",
        "accountability": "accountability",
    }
    return mapping.get(theme_raw.lower(), theme_raw.lower().replace(" ", "_"))


# ---------------------------------------------------------------------------
# DOCX text extraction
# ---------------------------------------------------------------------------
def extract_paragraphs(docx_path: Path) -> List[str]:
    """Returns list of non-empty paragraph strings from a .docx file."""
    try:
        from docx import Document
    except ImportError:
        raise ImportError("Run: pip install python-docx")

    doc = Document(docx_path)
    paragraphs = []
    for p in doc.paragraphs:
        text = p.text.strip()
        if text:
            paragraphs.append(text)
    return paragraphs


# ---------------------------------------------------------------------------
# Chunking
# ---------------------------------------------------------------------------
def chunk_paragraphs(paragraphs: List[str], max_words: int = MAX_WORDS_PER_CHUNK, overlap: int = OVERLAP_PARAGRAPHS) -> List[str]:
    """
    Groups paragraphs into chunks up to max_words.
    Carries the last `overlap` paragraphs from the previous chunk into the next.
    """
    if not paragraphs:
        return []

    chunks = []
    current = []
    current_words = 0

    for para in paragraphs:
        para_words = len(para.split())
        if current_words + para_words > max_words and current:
            chunks.append(" ".join(current))
            # Carry overlap paragraphs forward
            current = current[-overlap:] if overlap > 0 else []
            current_words = sum(len(p.split()) for p in current)
        current.append(para)
        current_words += para_words

    if current:
        chunks.append(" ".join(current))

    return chunks


# ---------------------------------------------------------------------------
# Embedding
# ---------------------------------------------------------------------------
def build_embedding_text(doc_name: str, theme_category: str, keywords: List[str], content: str) -> str:
    return "\n".join([
        f"Documento: {doc_name}",
        f"Categoría: {theme_category}",
        f"Palabras clave: {', '.join(keywords)}",
        f"Contenido: {content}",
    ])


def generate_embedding(text: str, api_key: str) -> Optional[List[float]]:
    try:
        response = requests.post(
            "https://api.openai.com/v1/embeddings",
            headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
            json={"model": "text-embedding-ada-002", "input": text.strip()[:8000]},
            timeout=30,
        )
        response.raise_for_status()
        return response.json()["data"][0]["embedding"]
    except Exception as e:
        logger.error(f"Embedding error: {e}")
        return None


# ---------------------------------------------------------------------------
# MongoDB helpers
# ---------------------------------------------------------------------------
def already_ingested(col, document_name: str) -> bool:
    return col.count_documents({"document_name": document_name}) > 0


# ---------------------------------------------------------------------------
# Main ingestion
# ---------------------------------------------------------------------------
def ingest_document(
    col,
    docx_path: Path,
    meta: Dict,
    api_key: str,
) -> int:
    """Ingests one .docx file. Returns number of chunks inserted."""
    doc_name = docx_path.stem  # filename without .docx

    if already_ingested(col, doc_name):
        logger.info(f"⏭️  Skipping (already ingested): {doc_name}")
        return 0

    logger.info(f"📄 Processing: {doc_name}")

    paragraphs = extract_paragraphs(docx_path)
    if not paragraphs:
        logger.warning(f"⚠️  No text found in {doc_name}")
        return 0

    chunks = chunk_paragraphs(paragraphs)
    logger.info(f"   → {len(paragraphs)} paragraphs → {len(chunks)} chunks")

    documents_to_insert = []
    for idx, chunk_text in enumerate(chunks):
        embedding_text = build_embedding_text(
            doc_name=doc_name,
            theme_category=meta["theme_category"],
            keywords=meta["keywords"],
            content=chunk_text,
        )

        embedding = generate_embedding(embedding_text, api_key)
        if embedding is None:
            logger.error(f"   ❌ Failed embedding for chunk {idx} of {doc_name}")
            continue

        chunk_hash = hashlib.md5(chunk_text.encode()).hexdigest()

        doc = {
            # Content
            "content": chunk_text,
            "chunk_index": idx,
            "total_chunks": len(chunks),
            # Source identity
            "document_name": doc_name,
            "document_file": docx_path.name,
            "source_category": "conocimiento_general",
            "knowledge_type": meta["knowledge_type"],
            # Classification
            "theme_category": meta["theme_category"],
            "language": meta["language"],
            "keywords": meta["keywords"],
            "content_summary": meta["content_summary"],
            # Embedding
            "embedding": embedding,
            "embedding_model": "text-embedding-ada-002",
            "embedding_text": embedding_text,
            "chunk_hash": chunk_hash,
            # Timestamps
            "ingested_at": datetime.utcnow(),
        }
        documents_to_insert.append(doc)
        time.sleep(EMBEDDING_DELAY)

    if documents_to_insert:
        col.insert_many(documents_to_insert)
        logger.info(f"   ✅ Inserted {len(documents_to_insert)} chunks for {doc_name}")

    return len(documents_to_insert)


def main():
    # Validate env
    api_key = os.getenv("OPENAI_API_KEY")
    mongo_uri = os.getenv("MONGODB_URI")
    db_name = os.getenv("MONGODB_DB_NAME", "puddle")

    if not api_key:
        raise ValueError("OPENAI_API_KEY not set")
    if not mongo_uri:
        raise ValueError("MONGODB_URI not set")

    # Connect
    client = MongoClient(mongo_uri)
    col = client[db_name][TARGET_COLLECTION]
    logger.info(f"🔗 Connected to {db_name}.{TARGET_COLLECTION}")

    # Load Excel metadata
    meta_index = load_metadata_index(EXCEL_PATH)

    # Find .docx files
    docx_files = sorted(DOCS_DIR.glob("*.docx"))
    logger.info(f"📁 Found {len(docx_files)} .docx files in {DOCS_DIR}")

    total_inserted = 0
    skipped = 0
    no_meta = []

    for docx_path in docx_files:
        doc_name = docx_path.stem
        meta = meta_index.get(doc_name)

        if meta is None:
            logger.warning(f"⚠️  No metadata found for {doc_name} — ingesting with defaults")
            meta = {
                "knowledge_type": "general",
                "country": "",
                "program": "",
                "theme_category": "unknown",
                "language": "es",
                "keywords": [],
                "content_summary": "",
                "status": "Available",
            }
            no_meta.append(doc_name)

        if meta.get("status", "Available").lower() not in ("available", ""):
            logger.info(f"⏭️  Skipping (status={meta['status']}): {doc_name}")
            skipped += 1
            continue

        inserted = ingest_document(col, docx_path, meta, api_key)
        total_inserted += inserted

    # Summary
    logger.info("\n" + "=" * 50)
    logger.info(f"✅ Ingestion complete")
    logger.info(f"   Total chunks inserted: {total_inserted}")
    logger.info(f"   Skipped (status): {skipped}")
    if no_meta:
        logger.info(f"   Missing Excel metadata: {no_meta}")

    # Quick verification
    logger.info(f"\n📊 Collection stats:")
    logger.info(f"   Total chunks: {col.count_documents({})}")
    logger.info(f"   Unique documents: {len(col.distinct('document_name'))}")
    logger.info(f"   Theme categories: {col.distinct('theme_category')}")

    client.close()


if __name__ == "__main__":
    main()
